from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
import argparse
import sys

"""
Collect command-line arguments
"""
def cli():
    parser = argparse.ArgumentParser()
    parser.add_argument('--file', '-f', help='your xls/xlsx source file', required=True)
    parser.add_argument('--sheet', '-s', help='Sheet name', required=True)
    parser.add_argument('--output', '-o', help='Book name for AVG graph result', default='default', required=False)
    return parser.parse_args()

def loadFile (add):
    # Funcion para cargar el archivo
    try:
        wb = load_workbook(add)
        return (wb)
    except Exception as e:
         sys.exit("No se pudo ejecutar el programa")
    raise

def loadSheet(wb, sht):
    # Funcion para cargar la hoja de calculo
    ws = wb.get_sheet_by_name(sht)
    return (ws)

def compAVG(ws):
    # Funcion para calcular el promedio de los datos de la columna 2
    try:
            x = ws.max_row
            y = ws.max_column

            sum = 0

            for i in range(y,x):
                sum += ws.cell(row=i, column=y).value
                pass

            avg = (sum / (x - 1))

            ws['A' + str(x+1)] = 'El promedio es'
            ws['B' + str(x+1)] = avg
    except Exception as e:
        sys.exit("No se pudo ejecutar el programa")
        raise

def generateChart(ws, wb, name):
    #D efine los parametros de la grafica y donde se ubiara
    try:
        chart = ScatterChart()
        data1 = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=17)
        data2 = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=17)
        series = Series(data2, data1)
        chart.series.append(series)
        ws.add_chart(chart, "E6")
        wb.save(name + ".xlsx") # Guarda la grafica en el nuevo documento creado
        print 'Su grafica ha sido generada satisfactoriamente. Por favor revise el archivo'
        pass
    except Exception as e:
        sys.exit("No se pudo ejecutar el programa")
        print ()
        raise

if __name__ == '__main__':
    args = cli()
    print(args)
    book = loadFile(args.file)
    sheet = loadSheet(book, args.sheet)
    compAVG(sheet)
    generateChart(sheet, book, args.output)
